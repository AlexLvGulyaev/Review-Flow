"""Export admin reports to CSV, XLSX, PDF (Unicode / Cyrillic)."""

from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Any

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.schemas.reports import BusinessProblemsReport, ChQualityReport, CustomerReviewsReport
from app.services.pdf_fonts import FONT_BOLD, FONT_REGULAR, register_cyrillic_fonts

try:
    from openpyxl import Workbook
except ImportError:  # pragma: no cover
    Workbook = None


def _period_label(period_meta) -> str:
    return f"{period_meta.date_from.date()} — {period_meta.date_to.date()} ({period_meta.period})"


def _write_table_rows(writer, title: str, rows, headers=None) -> None:
    writer.writerow([])
    if not rows:
        writer.writerow(headers or [title, "Количество", "Доля, %"])
        return
    if headers is not None:
        writer.writerow(headers)
    elif hasattr(rows[0], "label"):
        writer.writerow([title, "Количество", "Доля, %"])
    elif isinstance(rows[0], (list, tuple)) and len(rows[0]) == 2:
        writer.writerow([title, "Количество"])
    else:
        writer.writerow([title, "Количество", "Доля, %"])
    for row in rows:
        if hasattr(row, "label"):
            writer.writerow([row.label, row.count, row.share])
        else:
            writer.writerow(list(row))


def _xlsx_block(ws, title: str, rows, headers=None) -> None:
    ws.append([])
    ws.append(headers or [title, "Количество", "Доля, %"])
    for row in rows:
        if hasattr(row, "label"):
            ws.append([row.label, row.count, row.share])
        elif isinstance(row, (list, tuple)):
            ws.append(list(row))
        else:
            ws.append([row.label, row.count])


def _pdf_styles():
    register_cyrillic_fonts()
    return {
        "title": ParagraphStyle(
            "RepTitle",
            fontName=FONT_BOLD,
            fontSize=14,
            leading=18,
            spaceAfter=6,
        ),
        "heading": ParagraphStyle(
            "RepHeading",
            fontName=FONT_BOLD,
            fontSize=10,
            leading=13,
            spaceBefore=8,
            spaceAfter=4,
        ),
        "body": ParagraphStyle(
            "RepBody",
            fontName=FONT_REGULAR,
            fontSize=9,
            leading=12,
        ),
    }


def _pdf_table(data: list[list[str]]) -> Table:
    table = Table(data, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8ecf1")),
                ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
                ("FONTNAME", (0, 1), (-1, -1), FONT_REGULAR),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    return table


def _build_pdf(title: str, period_meta, sections: list[tuple[str, Any]], summary: str) -> bytes:
    buf = io.BytesIO()
    styles = _pdf_styles()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=1.5 * cm, leftMargin=1.5 * cm)
    story = [
        Paragraph(title, styles["title"]),
        Paragraph(_period_label(period_meta), styles["body"]),
        Spacer(1, 10),
    ]
    for heading, payload in sections:
        story.append(Paragraph(heading, styles["heading"]))
        if isinstance(payload, list) and payload and isinstance(payload[0], list):
            story.append(_pdf_table(payload))
        elif isinstance(payload, list) and payload and hasattr(payload[0], "label"):
            rows = [["Категория", "Кол-во", "Доля %"]] + [
                [r.label, str(r.count), str(r.share)] for r in payload
            ]
            story.append(_pdf_table(rows))
        elif isinstance(payload, list) and payload and len(payload[0]) == 2:
            rows = [["Показатель", "Значение"]] + [[str(a), str(b)] for a, b in payload]
            story.append(_pdf_table(rows))
        else:
            story.append(Paragraph(str(payload), styles["body"]))
        story.append(Spacer(1, 6))
    story.append(Paragraph("Сводка", styles["heading"]))
    story.append(Paragraph(summary or "—", styles["body"]))
    doc.build(story)
    return buf.getvalue()


def export_filename(report_key: str, period: str, ext: str) -> str:
    stamp = datetime.now().strftime("%Y%m%d")
    slug = {
        "customer-reviews": "customer-reviews",
        "business-problems": "biznes-svodka",
        "ch-quality": "ch-quality",
    }.get(report_key, report_key)
    return f"{slug}_{period}_{stamp}.{ext}"


# --- Customer reviews ---


def export_customer_reviews_csv(report: CustomerReviewsReport) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Отчёт: Обращения клиентов", _period_label(report.period)])
    w.writerow(["KPI", "Значение"])
    w.writerow(["Всего обращений", report.total_reviews])
    w.writerow(["Обработано", report.processed_reviews])
    w.writerow(["В работе", report.in_progress_reviews])
    w.writerow(["Средняя оценка", report.average_rating or ""])
    w.writerow(["Среднее время обработки, ч", report.average_processing_hours or ""])
    for title, items in [
        ("Обращения по дням", report.reviews_by_day),
        ("Продукты", report.by_product_area),
        ("Сценарии", report.by_scenario),
        ("Тональность", report.by_sentiment),
        ("Приоритет", report.by_priority),
    ]:
        w.writerow([])
        w.writerow([title, "Количество"])
        for item in items or []:
            w.writerow([item.label, item.count])
    _write_table_rows(w, "ТОП тем", report.top_topics)
    w.writerow([])
    w.writerow(["Сводка", report.summary])
    return buf.getvalue().encode("utf-8-sig")


def export_customer_reviews_xlsx(report: CustomerReviewsReport) -> bytes:
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed")
    wb = Workbook()
    ws = wb.active
    ws.title = "Обращения"
    ws.append(["Отчёт: Обращения клиентов", _period_label(report.period)])
    ws.append(["KPI", "Значение"])
    ws.append(["Всего обращений", report.total_reviews])
    ws.append(["Обработано", report.processed_reviews])
    ws.append(["В работе", report.in_progress_reviews])
    ws.append(["Средняя оценка", report.average_rating])
    ws.append(["Среднее время обработки, ч", report.average_processing_hours])
    for title, items in [
        ("Обращения по дням", report.reviews_by_day),
        ("Продукты", report.by_product_area),
        ("Сценарии", report.by_scenario),
        ("Тональность", report.by_sentiment),
        ("Приоритет", report.by_priority),
    ]:
        ws.append([])
        ws.append([title, "Количество"])
        for item in items or []:
            ws.append([item.label, item.count])
    _xlsx_block(ws, "ТОП тем", report.top_topics)
    ws.append(["Сводка", report.summary])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def export_customer_reviews_pdf(report: CustomerReviewsReport) -> bytes:
    sections = [
        (
            "KPI",
            [
                ["Всего обращений", report.total_reviews],
                ["Обработано", report.processed_reviews],
                ["В работе", report.in_progress_reviews],
                ["Средняя оценка", report.average_rating or "—"],
                ["Среднее время, ч", report.average_processing_hours or "—"],
            ],
        ),
    ]
    for title, items in [
        ("По дням", report.reviews_by_day),
        ("Продукты", report.by_product_area),
        ("Сценарии", report.by_scenario),
        ("Тональность", report.by_sentiment),
        ("Приоритет", report.by_priority),
    ]:
        if items:
            rows = [["Категория", "Кол-во"]] + [[i.label, str(i.count)] for i in items[:15]]
            sections.append((title, rows))
    if report.top_topics:
        sections.append(("ТОП тем", report.top_topics[:15]))
    return _build_pdf("Обращения клиентов", report.period, sections, report.summary)


# --- Business problems ---


def export_business_problems_csv(report: BusinessProblemsReport) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Отчёт: Бизнес-сводка", _period_label(report.period)])
    _write_table_rows(w, "ТОП жалоб", report.top_complaints)
    _write_table_rows(w, "ТОП предложений", report.top_suggestions)
    _write_table_rows(w, "ТОП благодарностей", report.top_gratitude)
    _write_table_rows(w, "Новые темы", report.new_topics)
    w.writerow([])
    w.writerow(["Сводка", report.summary])
    return buf.getvalue().encode("utf-8-sig")


def export_business_problems_xlsx(report: BusinessProblemsReport) -> bytes:
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed")
    wb = Workbook()
    ws = wb.active
    ws.title = "Бизнес-сводка"
    ws.append(["Отчёт: Бизнес-сводка", _period_label(report.period)])
    _xlsx_block(ws, "ТОП жалоб", report.top_complaints)
    _xlsx_block(ws, "ТОП предложений", report.top_suggestions)
    _xlsx_block(ws, "ТОП благодарностей", report.top_gratitude)
    _xlsx_block(ws, "Новые темы", report.new_topics)
    ws.append(["Сводка", report.summary])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def export_business_problems_pdf(report: BusinessProblemsReport) -> bytes:
    sections = [
        ("ТОП жалоб (сценарий «Жалоба»)", report.top_complaints),
        ("ТОП предложений (сценарий «Предложение»)", report.top_suggestions),
        ("ТОП благодарностей (сценарий «Благодарность»)", report.top_gratitude),
        ("Новые темы (кандидаты / новые ТС)", report.new_topics),
    ]
    return _build_pdf("Бизнес-сводка", report.period, sections, report.summary)


# --- CH quality ---


def export_ch_quality_csv(report: ChQualityReport) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Отчёт: Качество Controlled Hybrid", _period_label(report.period)])
    w.writerow(["Показатель", "Значение"])
    w.writerow(["Покрытие базы знаний, %", report.coverage_pct])
    w.writerow(["Доля ручных исправлений, %", report.override_rate_pct])
    w.writerow(["Низкая уверенность системы, %", report.low_confidence_rate_pct])
    w.writerow(["Создано типовых ситуаций", report.new_cases])
    w.writerow(["Добавлено retrieval-примеров", report.new_examples])
    w.writerow(["Кандидаты на расширение базы знаний", report.candidates_created])
    _write_table_rows(w, report.problematic_cases_title, report.problematic_cases)
    w.writerow([])
    w.writerow(["Сводка", report.summary])
    return buf.getvalue().encode("utf-8-sig")


def export_ch_quality_xlsx(report: ChQualityReport) -> bytes:
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed")
    wb = Workbook()
    ws = wb.active
    ws.title = "CH Quality"
    ws.append(["Отчёт: Качество Controlled Hybrid", _period_label(report.period)])
    ws.append(["Покрытие базы знаний, %", report.coverage_pct])
    ws.append(["Доля ручных исправлений, %", report.override_rate_pct])
    ws.append(["Низкая уверенность системы, %", report.low_confidence_rate_pct])
    ws.append(["Создано типовых ситуаций", report.new_cases])
    ws.append(["Добавлено retrieval-примеров", report.new_examples])
    ws.append(["Кандидаты на расширение базы", report.candidates_created])
    _xlsx_block(ws, report.problematic_cases_title, report.problematic_cases)
    ws.append(["Сводка", report.summary])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def export_ch_quality_pdf(report: ChQualityReport) -> bytes:
    sections = [
        (
            "Показатели",
            [
                ["Покрытие базы знаний, %", report.coverage_pct],
                ["Доля ручных исправлений, %", report.override_rate_pct],
                ["Низкая уверенность, %", report.low_confidence_rate_pct],
                ["Создано типовых ситуаций", report.new_cases],
                ["Добавлено retrieval-примеров", report.new_examples],
                ["Кандидаты на расширение базы", report.candidates_created],
            ],
        ),
        (report.problematic_cases_title, report.problematic_cases),
    ]
    summary = f"{report.problematic_cases_criterion} {report.summary}"
    return _build_pdf("Качество Controlled Hybrid", report.period, sections, summary)


EXPORT_HANDLERS = {
    "customer-reviews": {
        "csv": export_customer_reviews_csv,
        "xlsx": export_customer_reviews_xlsx,
        "pdf": export_customer_reviews_pdf,
    },
    "business-problems": {
        "csv": export_business_problems_csv,
        "xlsx": export_business_problems_xlsx,
        "pdf": export_business_problems_pdf,
    },
    "ch-quality": {
        "csv": export_ch_quality_csv,
        "xlsx": export_ch_quality_xlsx,
        "pdf": export_ch_quality_pdf,
    },
}

MEDIA_TYPES = {
    "csv": "text/csv; charset=utf-8",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}
